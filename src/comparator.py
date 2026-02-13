
import pandas as pd
import numpy as np

class Comparator:
    def __init__(self, db_manager):
        self.db = db_manager

    def get_all_transactions(self, table_name):
        return pd.read_sql(f"SELECT * FROM {table_name}", self.db.conn)

    def compare_transactions(self, date_field='data_ocorrencia', start_date=None, end_date=None):
        # Load data
        df_api = self.get_all_transactions('api_transactions')
        df_geral = self.get_all_transactions('geral_transactions')
        
        if df_api.empty and df_geral.empty:
            print("Nenhum dado encontrado para comparar.")
            return {}
            
        # Optional: Filter for financial transactions (valor_pago > 0)
        print(f"Filtrando registros com valor pago > 0 para comparação...")
        df_api = df_api[df_api['valor_pago'] > 0].copy()
        df_geral = df_geral[df_geral['valor_pago'] > 0].copy()
        
        # Filter by Date Range if provided
        if start_date and end_date:
            print(f"Filtrando por {date_field} entre {start_date} e {end_date}...")
            # Ensure format matches DB (DDMMYYYY)
            # Assuming inputs are YYYY-MM-DD or DDMMYYYY depending on GUI
            # Let's assume standard string comparison works if format is consistent
            # The parser stores as DDMMYYYY. 
            # We need to convert to datetime for correct comparison
            
            def filter_date(df, field, start, end):
                try:
                    # Convert DDMMYYYY to datetime
                    dt_series = pd.to_datetime(df[field], format='%d%m%Y', errors='coerce')
                    start_dt = pd.to_datetime(start, format='%d/%m/%Y')
                    end_dt = pd.to_datetime(end, format='%d/%m/%Y')
                    return df[(dt_series >= start_dt) & (dt_series <= end_dt)]
                except Exception as e:
                    print(f"Erro ao filtrar datas: {e}")
                    return df

            df_api = filter_date(df_api, date_field, start_date, end_date)
            df_geral = filter_date(df_geral, date_field, start_date, end_date)

        # Deduplicate
        print("Removendo duplicatas...")
        subset_cols = ['nosso_numero', 'valor_pago', 'data_ocorrencia']
        df_api = df_api.drop_duplicates(subset=subset_cols)
        df_geral = df_geral.drop_duplicates(subset=subset_cols)
        
        # Prepare key columns
        df_api['nosso_numero'] = df_api['nosso_numero'].astype(str).str.strip()
        df_geral['nosso_numero'] = df_geral['nosso_numero'].astype(str).str.strip()
        
        # Round monetary values to 2 decimal places to avoid float precision issues
        monetary_cols = ['valor_pago', 'valor_titulo', 'valor_liquido', 'juros_multa', 'desconto', 'abatimento', 'iof', 'outras_despesas', 'outros_creditos']
        for col in monetary_cols:
            if col in df_api.columns:
                df_api[col] = df_api[col].astype(float).round(2)
            if col in df_geral.columns:
                df_geral[col] = df_geral[col].astype(float).round(2)

        # Merge
        # We use 'nosso_numero' as the key. 
        # Note: If 'nosso_numero' is not unique, this merge will create a Cartesian product for duplicates.
        # Ideally we should deduplicate or use more keys (like 'numero_documento' or 'vencimento').
        # But 'nosso_numero' is the standard key.
        
        merged = pd.merge(
            df_api, 
            df_geral, 
            on='nosso_numero', 
            how='outer', 
            suffixes=('_api', '_geral'),
            indicator=True
        )
        
        # Split results
        missing_in_geral = merged[merged['_merge'] == 'left_only'].copy()
        missing_in_api = merged[merged['_merge'] == 'right_only'].copy()
        present_in_both = merged[merged['_merge'] == 'both'].copy()
        
        # Identify discrepancies
        # Compare important fields
        # value discrepancies
        present_in_both['diff_valor_pago'] = np.abs(present_in_both['valor_pago_api'] - present_in_both['valor_pago_geral'])
        
        # Tolerance for float comparison (e.g. 0.01)
        # Use selected date_field for discrepancy check
        date_field_api = f"{date_field}_api"
        date_field_geral = f"{date_field}_geral"
        
        discrepancies = present_in_both[
            (present_in_both['diff_valor_pago'] > 0.01) |
            (present_in_both[date_field_api] != present_in_both[date_field_geral])
        ].copy()
        
        return {
            'missing_in_geral': missing_in_geral,
            'missing_in_api': missing_in_api,
            'discrepancies': discrepancies,
            'matched': present_in_both,
            'all_data': merged
        }

    def generate_report(self, results, output_dir='reports'):
        import os
        from openpyxl.styles import Font, NamedStyle
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        if not results:
            print("Sem resultados para gerar relatório.")
            return

        # Helper to add total row
        def add_total_row(df, label_col='nosso_numero'):
            if df.empty: return df
            
            # Identify numeric columns
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            
            # Calculate totals
            total_row = {col: df[col].sum() for col in numeric_cols}
            total_row[label_col] = 'TOTAL'
            
            # Create DataFrame for total row and append
            # Using concat instead of append (deprecated)
            total_df = pd.DataFrame([total_row])
            return pd.concat([df, total_df], ignore_index=True)

        # Helper to apply formatting to a sheet
        def format_sheet(writer, sheet_name, df):
            try:
                worksheet = writer.sheets[sheet_name]
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # Apply number format to numeric columns
                # Iterate over columns in DataFrame to find numeric ones
                for i, col in enumerate(df.columns, 1): # 1-based index for openpyxl
                    try:
                        # Check if column is numeric
                        is_numeric = pd.api.types.is_numeric_dtype(df[col])
                    except:
                        is_numeric = False
                        
                    if is_numeric:
                        # Apply to all rows in this column (except header)
                        for row in range(2, max_row + 1):
                            cell = worksheet.cell(row=row, column=i)
                            # Only apply if value is number
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                
                # Bold the last row (TOTAL)
                # Assuming the last row is the total row we added
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=max_row, column=col)
                    cell.font = Font(bold=True)
                    
            except Exception as e:
                print(f"Erro ao formatar planilha {sheet_name}: {e}")

        # Summary
        summary = []
        summary.append("Relatório de Comparação CNAB240")
        summary.append("===============================")
        summary.append(f"Registros encontrados somente na API: {len(results['missing_in_geral'])}")
        summary.append(f"Registros encontrados somente no Arquivo Geral: {len(results['missing_in_api'])}")
        summary.append(f"Registros correspondentes (Total): {len(results['matched'])}")
        summary.append(f"  - Com divergências (Valor/Data): {len(results['discrepancies'])}")
        summary.append(f"  - Sem divergências: {len(results['matched']) - len(results['discrepancies'])}")
        
        with open(f'{output_dir}/resumo.txt', 'w', encoding='utf-8') as f:
            f.write('\n'.join(summary))
            
        # Excel Report
        # Select columns for cleaner report
        # For 'missing' tables, we want the columns from that side.
        
        output_file = f'{output_dir}/relatorio_completo_v2.xlsx'
        try:
            with pd.ExcelWriter(output_file) as writer:
                # Sheet 1: Divergências
                if not results['discrepancies'].empty:
                    # Determine date columns based on available data or default
                    date_col_api = 'data_ocorrencia_api' if 'data_ocorrencia_api' in results['discrepancies'].columns else 'data_credito_api'
                    date_col_geral = 'data_ocorrencia_geral' if 'data_ocorrencia_geral' in results['discrepancies'].columns else 'data_credito_geral'
                    
                    cols = ['nosso_numero', 'valor_pago_api', 'valor_pago_geral', 'diff_valor_pago', date_col_api, date_col_geral, 'file_source_api', 'file_source_geral']
                    # Add extra info like Nome Pagador if available
                    if 'nome_pagador_api' in results['discrepancies'].columns:
                        cols.insert(1, 'nome_pagador_api')
                    
                    # Filter cols that exist
                    valid_cols = [c for c in cols if c in results['discrepancies'].columns]
                    df_disc = results['discrepancies'][valid_cols].copy()
                    df_disc = add_total_row(df_disc)
                    df_disc.to_excel(writer, sheet_name='Divergencias', index=False)
                    format_sheet(writer, 'Divergencias', df_disc)
                
                # Sheet 2: Faltando no Geral
                if not results['missing_in_geral'].empty:
                    # Keep API columns
                    cols_api = [c for c in results['missing_in_geral'].columns if c.endswith('_api') or c == 'nosso_numero']
                    df_miss_geral = results['missing_in_geral'][cols_api].copy()
                    df_miss_geral = add_total_row(df_miss_geral)
                    df_miss_geral.to_excel(writer, sheet_name='Sobra_API', index=False)
                    format_sheet(writer, 'Sobra_API', df_miss_geral)
                    
                # Sheet 3: Faltando na API
                if not results['missing_in_api'].empty:
                    # Keep General columns
                    cols_geral = [c for c in results['missing_in_api'].columns if c.endswith('_geral') or c == 'nosso_numero']
                    df_miss_api = results['missing_in_api'][cols_geral].copy()
                    df_miss_api = add_total_row(df_miss_api)
                    df_miss_api.to_excel(writer, sheet_name='Sobra_Geral', index=False)
                    format_sheet(writer, 'Sobra_Geral', df_miss_api)
                    
                # Sheet 4: Todos
                if not results['matched'].empty:
                    df_matched = results['matched'].copy()
                    df_matched = add_total_row(df_matched)
                    df_matched.to_excel(writer, sheet_name='Correspondentes', index=False)
                    format_sheet(writer, 'Correspondentes', df_matched)
                else:
                    # Create an empty sheet if nothing matches, to avoid IndexError
                    pd.DataFrame({'Info': ['Nenhum registro correspondente encontrado']}).to_excel(writer, sheet_name='Info', index=False)
            
            print(f"Relatórios gerados com sucesso em '{output_file}'")
        except PermissionError:
            print(f"Erro: O arquivo '{output_file}' está aberto. Por favor, feche-o e tente novamente.")
